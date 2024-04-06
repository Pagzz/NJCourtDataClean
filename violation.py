import openpyxl
import logging
from openpyxl import Workbook

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def read_names_and_data_from_cleaned_data(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.worksheets[-1]  # Assuming the last sheet is the newest
    data_dict = {}  # Use a dictionary to map normalized (first_name, last_name) to their corresponding data
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Normalize and adjust the indices if needed
        key = (row[1].strip().lower(), row[0].strip().lower()) if row[1] and row[0] else ('', '')  # Normalize the key
        data = {
            'Issue_Date': row[3] if row[3] is not None else '',
            'Court': row[11] if row[11] is not None else '',
            'Violation': row[12] if row[12] is not None else '',
        }
        data_dict[key] = data
    logging.info(f"Data read from cleaned data: {len(data_dict)} records.")
    return data_dict

def find_matches_and_write_to_numbers(names_filepath, data_dict, output_filepath='matched_records.xlsx'):
    names_workbook = openpyxl.load_workbook(names_filepath)
    names_sheet = names_workbook.worksheets[-1]  # Assuming the relevant sheet is the last one
    
    wb = Workbook()
    ws = wb.active
    # Update headers to include 'CellPhone'
    headers = ['First_Name', 'Last_Name', 'Issue_Date', 'Court', 'Violation', 'CellPhone']
    ws.append(headers)
    
    match_count = 0
    for row in names_sheet.iter_rows(min_row=2, values_only=True):
        key = (row[1].strip().lower(), row[0].strip().lower()) if row[1] and row[0] else ('', '')  # Normalize
        cell_phone = row[2] if len(row) > 2 else ''
        matched_data = data_dict.get(key, {'Issue_Date': '', 'Court': '', 'Violation': ''})
        ws.append([row[1], row[0]] + [matched_data[header] for header in ['Issue_Date', 'Court', 'Violation']] + [cell_phone])
        match_count += 1 if key in data_dict else 0
    logging.info(f"Matched records written to {output_filepath}. Total matches found: {match_count}")
    wb.save(output_filepath)

if __name__ == '__main__':
    cleaned_data_filepath = 'C:\\Users\\rp122\\OneDrive\\Desktop\\dataUSA\\Excel\\cleaned_data.xlsx'
    names_filepath = 'C:\\Users\\rp122\\OneDrive\\Desktop\\dataUSA\\Excel\\numbers.xlsx'
    output_filepath = 'C:\\Users\\rp122\\OneDrive\\Desktop\\dataUSA\\Excel\\Match.xlsx'
    
    data_dict = read_names_and_data_from_cleaned_data(cleaned_data_filepath)
    find_matches_and_write_to_numbers(names_filepath, data_dict, output_filepath)
