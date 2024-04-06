import aiohttp
import asyncio
import json
import openpyxl
from openpyxl import Workbook
import logging
import re

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Toggle deduplication process
ENABLE_DEDUPLICATION = True  # Set to False to disable deduplication


async def get_auth_token():
    async with aiohttp.ClientSession() as session:
        url = 'http://www.datairis.co/V1/auth/subscriber/'
        headers = {
            'SubscriberUsername': 'mintmediaapp',
            'SubscriberPassword': 'FNCcKWyi0bAXnoPAka8L',
            'AccountUsername': 'mintmediaact',
            'AccountPassword': 'reference',
            'AccountDetailsRequired': 'true',
            'SubscriberID': '224',
        }
        params = {
            'AccessToken': 'ad381-a59ec-4d7d',
        }
        try:
            async with session.get(url, headers=headers, params=params) as response:
                if response.status == 200:
                    data = await response.json()
                    logging.info("Token obtained successfully.")
                    return data['Response']['responseDetails']['TokenID']
                else:
                    logging.error(f"Failed to get auth token, HTTP status: {response.status}")
                    return None
        except Exception as e:
            logging.error(f"Error obtaining auth token: {e}")
            return None

async def add_all_search_criteria(token_id, database_type, criteria):
    url = f"http://www.datairis.co/V1/criteria/search/addall/{database_type}"
    headers = {
        'TokenID': token_id,
        'Content-Type': 'application/json'
    }
    payload = json.dumps(criteria)

    try:
        async with aiohttp.ClientSession() as session:
            async with session.put(url, headers=headers, data=payload) as response:
                response_data = await response.json()
                logging.info(f"Criteria add response: {response_data}")
                if response.status == 200:
                    return response_data
                else:
                    logging.error(f"Failed to add search criteria, HTTP status: {response.status}")
                    return None
    except Exception as e:
        logging.error(f"Error adding search criteria: {e}")
        return None


async def get_search_results(token_id, database_type, start, end):
    url = f"http://www.datairis.co/V1/search/{database_type}/?Start={start}&End={end}"
    headers = {
        'TokenID': token_id
    }

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as response:
                response_data = await response.json()
                logging.info(f"Search results: {response_data}")
                if response.status == 200:
                    return response_data
                else:
                    logging.error(f"Failed to get search results, HTTP status: {response.status}")
                    return None
    except Exception as e:
        logging.error(f"Error fetching search results: {e}")
        return None
    
    
def read_criteria_from_excel(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.worksheets[-1]  # Assuming the last sheet is the newest
    criteria_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Ensure you're pulling from the correct columns; adjust indices as necessary
        last_name = row[0] if row[0] is not None else ''
        first_name = row[1] if row[1] is not None else ''
        #middle_int = row[2] if row [2] is not None else ''
        #Physical_Address = row[6] if row[6] is not None else ''
        Physical_City = row[7] if row[7] is not None else ''
        Physical_State = row[8] if row[8] is not None else ''
        #Physical_Zip = row[9] if row[9] is not None else ''
       
        criteria = {
            'Last_Name': last_name, 
            'First_Name': first_name,
            #'Physical_Address': Physical_Address,
            'Physical_City':  Physical_City,
            'Physical_State': Physical_State,
            #'Middle_Initial': middle_int,
            #'Physical_Zip': Physical_Zip,

        }
        criteria_list.append(criteria)
    logging.info(f"Read {len(criteria_list)} criteria from Excel.")
    return criteria_list

def clean_html_tags(text):
    """Remove HTML tags from a string."""
    clean_text = re.sub(r'<[^>]+>', '', text)  # Regular expression to match HTML tags and replace them with an empty string
    return clean_text

def write_results_to_excel(results, filepath='API_Search_Results.xlsx'):
    wb = Workbook()
    ws = wb.active
    headers = [
        'Last_Name', 'First_Name', 'CellPhone', 'Scrubbed_Phoneable', 'Issue_Date', 'Court', 'Violation'  
    ]
    ws.append(headers)
    for result in results:
        # Ensure any necessary HTML cleaning or data preprocessing is applied
        if 'First_Name' in result:
            result['First_Name'] = clean_html_tags(result['First_Name'])
        # Populate the row with data for each header, handling missing data gracefully
        row = [result.get(header, '') for header in headers]
        ws.append(row)
    
    logging.info(f"Writing to Excel file at: {filepath}")
    wb.save(filepath)

async def main():
    token_id = await get_auth_token()
    if not token_id:
        logging.error("Failed to obtain token, terminating script.")
        return

    # The cleaned data filepath remains unchanged
    filepath = 'C:\\Users\\rp122\\OneDrive\\Desktop\\dataUSA\\Excel\\cleaned_data.xlsx'
    
    # Now using the updated function for reading criteria and additional data
    data_list = read_criteria_from_excel(filepath)
    if not data_list:
        logging.warning("Data list is empty, nothing to process.")
        return

    all_results = []

    # This loop is retained from your original script; assuming it's necessary for your application logic
    for criteria in data_list:
        
        response = await add_all_search_criteria(token_id, 'consumer', criteria)
        if not response:
            logging.error("Failed to add search criteria for a record, skipping.")
            continue

        start = 1
        end = 10
        results = await get_search_results(token_id, 'consumer', start, end)
        if results:
            search_results = results.get('Response', {}).get('responseDetails', {}).get('SearchResult', {}).get('searchResultRecord', [])
            for item in search_results:
                record = {field['fieldID']: field['fieldValue'] for field in item.get('resultFields', [])}
                # Default 'Scrubbed_Phoneable' to 'Not Available' if not found
                if 'Scrubbed_Phoneable' not in record:
                    record['Scrubbed_Phoneable'] = 'Not Available'
                all_results.append(record)
    # Assume deduplication and other processing as in your original script
    if ENABLE_DEDUPLICATION:
        unique_results = []
        seen = set()
        for result in all_results:
            identifier = result.get('CellPhone')  # or another unique identifier
            if identifier and identifier not in seen:
                unique_results.append(result)
                seen.add(identifier)
        logging.info(f"Prepared to write results, total unique records: {len(unique_results)}")
    else:
        unique_results = all_results
        logging.info(f"Deduplication disabled, writing all {len(all_results)} results.")

    if not unique_results:
        logging.warning("No results to write to Excel.")
    else:
        # Updated output filepath remains the same
        output_filepath = 'C:\\Users\\rp122\\OneDrive\\Desktop\\dataUSA\\Excel\\numbers.xlsx'
        
        # Writing the results, including new fields, to Excel
        write_results_to_excel(unique_results, output_filepath)
        logging.info("Process completed, results written to Excel.")

if __name__ == '__main__':
    asyncio.run(main())